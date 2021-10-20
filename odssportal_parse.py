from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, ElementNotInteractableException
import openpyxl
import time


def wait_until_loadelem_become_invis(seconds):
    time.sleep(0.5)
    WebDriverWait(browser, seconds).until(EC.invisibility_of_element_located((By.ID, "event-wait-msg")))
    browser.execute_script('window.stop();')


def login():
     browser.get('https://www.oddsportal.com/login/')
     wait_until_loadelem_become_invis(30)
     log = browser.find_element_by_id('login-username1')
     log.send_keys('some_login')
     pas = browser.find_element_by_id('login-password1')
     pas.send_keys('some_pswrd')
     button = browser.find_element_by_xpath('//button[@class="inline-btn-2"]')
     button.click()
     wait_until_loadelem_become_invis(30)
     time.sleep(1)


def collect_refs(link_next, page):
    try:
        while page < 9:
            browser.get(link_next)
            wait_until_loadelem_become_invis(30)
            ref_elems = browser.find_elements_by_xpath('//td[@class="name table-participant"]/a')
            for elem in ref_elems:
                refs.append(elem.get_attribute('href'))
            page += 1
            link_next = link + '#/page/' + str(page) + '/'
    except NoSuchElementException or StaleElementReferenceException:
        return None


def collect_statistics(refs, statistics, odds_dict):
    key_number = len(refs)
    for ref in refs:
        browser.get(ref)
        collect_matchinfo_and_matchscore(statistics)
        statistics = sum([i.split(' - ') for i in statistics], [])
        statistics[2], statistics[3] = statistics[3], statistics[2]
        statistics = sum([i.split(':') for i in statistics], [])
        # collect_handicap0_odds(statistics)
        # collect_doublechanse_odds(statistics)
        # collect_asian_handicap(statistics)
        odds_dict.update({key_number: statistics.copy()})
        statistics.clear()
        print(key_number, end=' ')
        print(odds_dict[key_number])
        key_number -= 1


def collect_matchinfo_and_matchscore(statistics):
    wait_until_loadelem_become_invis(30)
    for selector in common_selectors:
        if common_selectors.index(selector) == 0:
            statistics.append(browser.find_element_by_xpath(selector).text[-18:-7])
        else:
            statistics.append(browser.find_element_by_xpath(selector).text)


def collect_handicap0_odds(statistics):
    browser.find_element_by_xpath('//a[@onmousedown="uid(6)._onClick();return false;"]').click()
    wait_until_loadelem_become_invis(30)
    # time.sleep(2)
    for selector in common_selectors[3:5]:
        statistics.append(browser.find_element_by_xpath(selector).text)


def collect_doublechanse_odds(statistics):
    try:
        browser.find_element_by_id('tab-sport-others').click()
        time.sleep(0.5)
        browser.find_element_by_xpath('//a[@onclick="uid(0).hideMore();uid(8)._onClick();return false;"]').click()
        wait_until_loadelem_become_invis(30)
        for selector in common_selectors[3:]:
            statistics.append(browser.find_element_by_xpath(selector).text)
    except NoSuchElementException or ElementNotInteractableException:
        browser.find_element_by_xpath('//a[@onmousedown="uid(8)._onClick();return false;"]').click()
        wait_until_loadelem_become_invis(30)
        for selector in common_selectors[3:]:
            statistics.append(browser.find_element_by_xpath(selector).text)
    # time.sleep(2)



def collect_asian_handicap(statistics):
    browser.find_element_by_xpath('//a[@onmousedown="uid(4)._onClick();return false;"]').click()
    wait_until_loadelem_become_invis(30)
    # time.sleep(2)
    for selector in handicap_selectors:
        try:
            statistics.append(browser.find_element_by_xpath(selector).text)
        except NoSuchElementException:
            statistics.append('')


t1 = time.time()
refs, statistics = [], []
odds_dict = {}
page = 1
# link = 'https://www.oddsportal.com/soccer/england/premier-league-2018-2019/results/'
link = 'https://www.oddsportal.com/soccer/england/premier-league-2004-2005/results/'
options = webdriver.ChromeOptions()
options.add_argument('headless')
caps = DesiredCapabilities.CHROME
browser = webdriver.Chrome(chrome_options=options, desired_capabilities=caps)
caps["pageLoadStrategy"] = "none"
'''caps = DesiredCapabilities.CHROME
browser = webdriver.Chrome(desired_capabilities=caps)
caps["pageLoadStrategy"] = "none"'''
date_selector = '//p[contains(@class, "date datet")]'
teams_selector = '//div[@id="col-content"]/h1'
score_selector = '//p[@class="result"]/strong'
first_selector = '//tr[@class="highest"]/td[@class="right"][1]'
second_selector = '//tr[@class="highest"]/td[@class="right"][2]'
third_selector = '//tr[@class="highest"]/td[@class="right"][3]'
common_selectors = [date_selector, teams_selector, score_selector, first_selector, second_selector, third_selector]
handicap_selectors = ['//a[contains(text(), "-2.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "-2.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "-1.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "-1.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "+1.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "+1.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "+2.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"]',
                      '//a[contains(text(), "+2.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"]']
browser.implicitly_wait(1)
collect_refs(link, page)
login()
collect_statistics(refs, statistics, odds_dict)

workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
for key, values in odds_dict.items():
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1
workbook.save(filename="odds.xlsx")
t2 = time.time()
print((t2-t1)/60)

# //tr[@class="highest"]/td[@class="right"]                                 # кэфы
# //div[@id="col-content"]/h1                                               # команды
# //div[@id="event-status"]//strong                                         # результат
# //a[@onmousedown="uid(6)._onClick();return false;"]                       # draw no bet
# //a[@onclick="uid(0).hideMore();uid(8)._onClick();return false;"]         # double chanse
# //a[@onmousedown="uid(4)._onClick();return false;"]                       # asian handicap
# page.togleTableContent('P-1.00-0-0',this);return false;
# //a[contains(text(), "+1")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"]    # левый гандикап +1
# //a[contains(text(), "+1.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"]  # правый гандикап +1.5
#  sasha121081
#  q87654321q

# //a[contains(text(), "-2.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"] # 1(-2.5)
# //a[contains(text(), "-2.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"] # 2(+2.5)
# //a[contains(text(), "-1.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"] # 1(-1.5)
# //a[contains(text(), "-1.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"] # 2(+1.5)
# //a[contains(text(), "+1.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"] # 1(+1.5)
# //a[contains(text(), "+1.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"] # 2(-1.5)
# //a[contains(text(), "+2.5")]/../../span[@class="avg chunk-odd nowrp"][2]/a[@xparam="odds_text"] # 1(+2.5)
# //a[contains(text(), "+2.5")]/../../span[@class="avg chunk-odd nowrp"][1]/a[@xparam="odds_text"] # 2(-2.5)




